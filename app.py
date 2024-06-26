import streamlit as st
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
import time
import openpyxl


def get_videos(url):
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    driver.get(url)
    time.sleep(5)
    for _ in range(10):
        driver.execute_script("window.scrollTo(0, document.documentElement.scrollHeight);")
        time.sleep(3)
    videos = driver.find_elements(By.XPATH, '//*[@id="video-title-link"]')
    video_data = {}
    for video in videos:
        title = video.get_attribute('title')
        url = video.get_attribute('href')
        video_data[title] = f'https://www.youtube.com{url}'
    driver.quit()
    return video_data


def add_links_to_excel(video_data, file_name):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "YouTube Links"
    sheet.cell(row=1, column=1, value="Video Title")
    sheet.cell(row=1, column=2, value="Url")
    
    for i, (title, url) in enumerate(video_data.items(), start=2):
        sheet.cell(row=i, column=1, value=title)
        sheet.cell(row=i, column=2, value=url)

    workbook.save(f"{file_name}.xlsx")
    return f"{file_name}.xlsx"


st.title("YouTube Channel Video Scraper")
st.write("Enter the URL of the YouTube channel's video page and get a list of video titles and URLs in an Excel file.")

url = st.text_input("YouTube Channel URL of Videos Section", "https://www.youtube.com/@nationalcareerservice-indi3451/videos")
file_name = st.text_input("Output Excel File Name", "Youtube_Urls_Titles")

if st.button("Get Video Links"):
    with st.spinner("Scraping video links..."):
        video_data = get_videos(url)
        file_path = add_links_to_excel(video_data, file_name)
        st.success(f"Links added to {file_path} successfully.")
        st.download_button(label="Download Excel file", data=open(file_path, "rb").read(), file_name=f"{file_name}.xlsx")
st.markdown("[Github](https://github.com/Ansumanbhujabal) | [LinkedIn](https://www.linkedin.com/in/ansuman-simanta-sekhar-bhujabala-30851922b/) | © 2024 Ansuman Bhujabala")
st.markdown("[Email- ansumanbhujabal1@gmail.com) ")
